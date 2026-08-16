#!/usr/bin/env python3
"""Build P4 HR data-collection Excel template (allowance rates + salary scales + FY2569 policy).

Usage (from python/): venv/Scripts/python.exe build_p4_template.py
Output: docs/templates/P4_ข้อมูลงบบุคลากร_2569.xlsx

- Allowance types are read live from the `allowance_types` table so the dropdown
  can never drift from the seeded catalog.
- All dropdown sources live on a hidden "Lists" sheet referenced by range —
  Excel caps inline DV strings at 255 chars, which the catalog list exceeds.
"""
from pathlib import Path

import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from db_config import get_db_config

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "templates" / "P4_ข้อมูลงบบุคลากร_2569.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SAMPLE_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill("solid", fgColor="E2EFDA")

# ขอบเขตระดับตำแหน่งของหน่วยงาน (ประเภททั่วไป) — ยืนยันโดยผู้ใช้ 2026-08-16
LEVELS = ["ปฏิบัติงาน", "ชำนาญงาน", "อาวุโส", "ทักษะพิเศษ"]
LEVELS_NULL = ["(ทุกระดับ)"] + LEVELS
# ตรง ENUM employee_category (migration 076 — "3 ค่าเท่านั้น ห้ามเพิ่ม")
CATEGORIES = ["ข้าราชการ", "พนักงานราชการ", "ลูกจ้างประจำ"]
# ตรง ENUM ใน migrations 087/088 และ frontend/src/lib/personnel.ts — แก้ต้องแก้พร้อมกัน
VACANCY_RULES = [
    "transfer_request | หนังสือขอรับโอน",
    "eligibility_list | ประกาศขึ้นบัญชี",
    "ready_to_fill | อัตราพร้อมบรรจุ",
]


def fetch_allowance_types():
    cfg = get_db_config()
    try:
        conn = mysql.connector.connect(
            host=cfg["host"], port=cfg["port"], user=cfg["user"],
            password=cfg["password"], database=cfg["database"], charset="utf8mb4",
        )
    except mysql.connector.Error as e:
        raise SystemExit(f"ต่อฐานข้อมูลไม่ได้ ({cfg['host']}:{cfg['port']}/{cfg['database']}) — เปิด MySQL ก่อน: {e}")
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT code, name_th FROM allowance_types "
            "WHERE is_active = 1 AND deleted_at IS NULL ORDER BY id"
        )
        rows = cursor.fetchall()
    finally:
        conn.close()
    if not rows:
        raise SystemExit("allowance_types table is empty - run migration 090 first")
    return [f"{r['code']} | {r['name_th']}" for r in rows]


def label(allowance_types, code):
    """Look a dropdown label up from the live catalog by code (never hardcode)."""
    for t in allowance_types:
        if t.split(" | ", 1)[0] == code:
            return t
    raise SystemExit(f"allowance type code not found in DB: {code}")


def build_lists_sheet(wb, columns):
    """Turn the default sheet into a hidden dropdown-source sheet; return range formulas."""
    ws = wb.active
    ws.title = "Lists"
    ws.sheet_state = "hidden"
    formulas = []
    for i, (name, values) in enumerate(columns, start=1):
        col = get_column_letter(i)
        ws.cell(row=1, column=i, value=name)
        for r, v in enumerate(values, start=2):
            ws.cell(row=r, column=i, value=v)
        formulas.append(f"Lists!${col}$2:${col}${len(values) + 1}")
    return formulas


def style_sheet(ws, headers, widths, row0_note=None):
    ws.freeze_panes = "A3"
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    if row0_note:
        ws.cell(row=1, column=1, value=row0_note).font = Font(italic=True, size=10, color="1F4E79")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))


def add_dropdown(ws, col_letter, formula, first_row=3, last_row=200):
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        errorTitle="ค่าไม่ถูกต้อง",
        error="กรุณาเลือกจากรายการ dropdown ที่ให้ไว้เท่านั้น",
        promptTitle="เลือกจากรายการ",
        prompt="คลิกที่ลูกศรเพื่อเลือกค่า",
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def write_samples(ws, rows, start_row=3):
    """เขียนแถวตัวอย่าง (สีเหลือง) — ใช้ร่วมทุกแผ่น"""
    for r, row in enumerate(rows, start=start_row):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = SAMPLE_FILL
            cell.alignment = Alignment(wrap_text=True)


def build_workbook(allowance_types):
    wb = Workbook()
    rng_type, rng_level_null, rng_level, rng_cat, rng_derived, rng_vacancy = build_lists_sheet(wb, [
        ("ชนิดเงินเพิ่ม", allowance_types),
        ("ระดับตำแหน่ง (รวมทุกระดับ)", LEVELS_NULL),
        ("ระดับตำแหน่ง", LEVELS),
        ("ประเภทบุคลากร", CATEGORIES),
        ("เท่ากับชนิดเงินเพิ่ม", ["-"] + [label(allowance_types, "POSITION_ALLOWANCE")]),
        ("เกณฑ์อัตราว่าง", VACANCY_RULES),
    ])

    # ---------- Sheet 1: อัตราเงินเพิ่ม (allowance_rates) ----------
    ws1 = wb.create_sheet("1_อัตราเงินเพิ่ม")
    h1 = [
        "ชนิดเงินเพิ่ม *",
        "ระดับตำแหน่ง",
        "สายงาน",
        "จำนวนเงิน (บาท/เดือน)",
        "ร้อยละของเงินเดือน (%)",
        "เท่ากับชนิดเงินเพิ่ม (derived)",
        "พื้นสำรอง (บาท)",
        "วันเริ่มมีผล *",
        "วันสิ้นสุด (ถ้ามี)",
        "เลขที่คำสั่ง/ประกาศ *",
        "หมายเหตุ",
    ]
    w1 = [34, 16, 18, 16, 14, 26, 14, 14, 14, 22, 30]
    style_sheet(
        ws1, h1, w1,
        "อัตราเงินเพิ่ม (allowance_rates) ปีงบประมาณ 2569 — กรอก 1 แถวต่อ 1 อัตรา · ช่อง * = ต้องกรอก · "
        "เว้นระดับ/สายงานว่าง = ทุกระดับ/ทุกสาย · แถวสีเหลืองเป็นตัวอย่าง ใส่ตัวเลขจริงแทนได้",
    )
    add_dropdown(ws1, "A", rng_type)
    add_dropdown(ws1, "B", rng_level_null)
    add_dropdown(ws1, "F", rng_derived)

    samples1 = [
        [label(allowance_types, "POSITION_ALLOWANCE"), "(ทุกระดับ)", "", 4200, "", "-", "", "2025-10-01", "", "ตัวอย่าง", ""],
        [label(allowance_types, "HOUSE_RENT"), "(ทุกระดับ)", "", 3000, "", "-", "", "2025-10-01", "", "ตัวอย่าง", "เป็นเพดานสูงสุด — จ่ายตามจริง ไม่บวกเต็มจำนวน"],
    ]
    write_samples(ws1, samples1)

    # ---------- Sheet 2: อัตราเงินเดือน (salary_scales) ----------
    ws2 = wb.create_sheet("2_อัตราเงินเดือน")
    h2 = [
        "ประเภทบุคลากร *",
        "ประเภทตำแหน่ง (ข้อมูลประกอบ)",
        "ระดับตำแหน่ง *",
        "เงินเดือนขั้นต่ำ (บาท) *",
        "เงินเดือนขั้นสูง (บาท) *",
        "วันเริ่มใช้ *",
        "เลขที่คำสั่ง/ประกาศ *",
        "หมายเหตุ",
    ]
    w2 = [18, 22, 18, 18, 18, 14, 22, 30]
    style_sheet(
        ws2, h2, w2,
        "อัตราเงินเดือนขั้นต่ำ–สูง (salary_scales) ปีงบประมาณ 2569 — กรอก 1 แถวต่อ 1 ระดับ · "
        "ขั้นสูงคือเพดานที่ใช้ตัดยอดประมาณการเลื่อนเงินเดือน · คอลัมน์ \"ประเภทตำแหน่ง\" เป็นข้อมูลประกอบ ไม่นำเข้าระบบ",
    )
    add_dropdown(ws2, "A", rng_cat)
    add_dropdown(ws2, "C", rng_level)

    samples2 = [
        ["ข้าราชการ", "ประเภททั่วไป", "ปฏิบัติงาน", 15000, 25000, "2025-10-01", "ตัวอย่าง", ""],
        ["ข้าราชการ", "ประเภททั่วไป", "ชำนาญงาน", 18000, 30000, "2025-10-01", "ตัวอย่าง", ""],
        ["ข้าราชการ", "ประเภททั่วไป", "อาวุโส", 22000, 38000, "2025-10-01", "ตัวอย่าง", ""],
        ["ข้าราชการ", "ประเภททั่วไป", "ทักษะพิเศษ", 25000, 42000, "2025-10-01", "ตัวอย่าง", ""],
    ]
    write_samples(ws2, samples2)

    # ---------- Sheet 3: นโยบายปี 2569 ----------
    ws3 = wb.create_sheet("3_นโยบายปี2569")
    h3 = ["รายการ", "ค่าที่ต้องระบุ *", "คำอธิบาย"]
    w3 = [30, 34, 60]
    style_sheet(ws3, h3, w3, "นโยบายการคำนวณงบบุคลากร ปีงบประมาณ 2569 (personnel_budget_policy)")

    policy_rows = [
        ["เกณฑ์อัตราว่างปี 2569 (vacancy_rule)", "", "อัตราว่างจะนับเข้างบเฉพาะที่ผ่านเกณฑ์นี้ — เลือกจาก dropdown"],
        ["เลขที่หนังสือเวียน/ประกาศ", "", "เลขที่หนังสือที่กำหนดเกณฑ์อัตราว่างปี 2569 (ข้อมูลอ้างอิง — ไม่นำเข้าระบบ)"],
        ["วันที่หนังสือ", "", "วันที่ของหนังสือเวียน/ประกาศ (ข้อมูลอ้างอิง — ไม่นำเข้าระบบ)"],
        ["โหมดคำนวณ (calc_mode)", "prorate", "กำหนดตายตัวแล้ว — ไม่ต้องแก้"],
        ["ช่องปรับ buffer (%)", "", "เว้นว่างได้ — ใช้เมื่อต้องการเผื่อช่องว่างของเงินเพิ่มแบบผูกบุคคล (เช่น ค่าเช่าบ้าน)"],
    ]
    for r, row in enumerate(policy_rows, start=3):
        for c, v in enumerate(row, start=1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if c == 2 and v == "prorate":
                cell.fill = NOTE_FILL
    add_dropdown(ws3, "B", rng_vacancy, first_row=3, last_row=3)

    # ---------- Sheet 4: คำแนะนำ ----------
    ws4 = wb.create_sheet("คำแนะนำ")
    tips = [
        "วิธีกรอกเทมเพลตนี้",
        "",
        "1) แผ่น \"1_อัตราเงินเพิ่ม\" — อัตราเงินเพิ่มทั้งหมดของปีงบ 2569 ตามรายการใน dropdown",
        "   - เลือกชนิดเงินเพิ่มจาก dropdown (คอลัมน์ A) และระดับตำแหน่งจาก dropdown (คอลัมน์ B)",
        "   - เว้นระดับ/สายงานว่าง = ใช้อัตราเดียวกันทุกตำแหน่ง",
        "   - กรอกจำนวนเงิน หรือ ร้อยละ อย่างใดอย่างหนึ่ง (เงินประจำตำแหน่งกรอกจำนวนเงิน)",
        "   - ค.ต.น. = เท่ากับเงินประจำตำแหน่งในระดับเดียวกัน — มีสิทธิ์เฉพาะระดับชำนาญการพิเศษขึ้นไป (ประเภทวิชาการ)",
        "     หน่วยที่มีเฉพาะประเภททั่วไปจึงไม่ต้องกรอก · หากต้องกรอกจริง: เลือก \"POSITION_ALLOWANCE\" ในคอลัมน์ F",
        "     และกรอกพื้นสำรอง (คอลัมน์ G) เฉพาะระดับที่มีพื้น — ว่าง = ไม่มีพื้น (ได้ 0)",
        "   - วันเริ่มมีผลใช้ 2025-10-01 (ต้นปีงบ 2569) เป็นหลัก ยกเว้นอัตราเริ่มใช้กลางปี",
        "",
        "2) แผ่น \"2_อัตราเงินเดือน\" — อัตราเงินเดือนขั้นต่ำ/ขั้นสูงของแต่ละระดับ",
        "   - ขั้นสูง (คอลัมน์ E) สำคัญมาก: ใช้เป็นเพดานตัดยอดประมาณการเลื่อน — ต้องถูกต้อง",
        "   - คอลัมน์ \"ประเภทตำแหน่ง\" เป็นข้อมูลประกอบเพื่ออ่านง่าย — ไม่นำเข้าระบบ",
        "",
        "3) แผ่น \"3_นโยบายปี2569\" — เกณฑ์อัตราว่างปี 2569 (แถวเลขที่/วันที่หนังสือเป็นข้อมูลอ้างอิง ไม่นำเข้าระบบ)",
        "",
        "4) กรอกเสร็จ ส่งไฟล์กลับมา ระบบจะนำเข้าโดยเจ้าหน้าที่ (แถวสีเหลืองเป็นตัวอย่าง — ลบหรือแทนที่ได้)",
        "",
        "ขอบเขตระดับตำแหน่งในเทมเพลตนี้: ประเภททั่วไป (ปฏิบัติงาน · ชำนาญงาน · อาวุโส · ทักษะพิเศษ)",
        "ถ้าหน่วยงานมีตำแหน่งประเภทอื่น (วิชาการ/อำนวยการ/บริหาร) แจ้งเจ้าหน้าที่ก่อนกรอกเพื่อเพิ่มรายการระดับ",
    ]
    ws4.column_dimensions["A"].width = 110
    for r, line in enumerate(tips, start=1):
        cell = ws4.cell(row=r, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True)
        if line.startswith(("1)", "2)", "3)", "4)")):
            cell.font = Font(bold=True)

    return wb


def main():
    allowance_types = fetch_allowance_types()
    wb = build_workbook(allowance_types)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
