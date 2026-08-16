#!/usr/bin/env python3
"""Import P4 HR data-collection workbook into the personnel-budget tables.

Reads the Excel template produced by build_p4_template.py (or the same file
filled in by HR) and imports it into:
  - allowance_rates          (sheet 1_อัตราเงินเพิ่ม)
  - salary_scales            (sheet 2_อัตราเงินเดือน)
  - personnel_budget_policies (sheet 3_นโยบายปี2569 — vacancy_rule + buffer)

Default mode is a dry run that only validates and prints what would change.
Use --apply to write to the database (single transaction).

Usage (from python/): venv/Scripts/python.exe import_p4_template.py [path.xlsx]
                      venv/Scripts/python.exe import_p4_template.py path.xlsx --apply
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# ---- ตรงกับ constants ใน build_p4_template.py (แก้พร้อมกัน) -----------------
LEVELS = ["ปฏิบัติงาน", "ชำนาญงาน", "อาวุโส", "ทักษะพิเศษ"]
LEVELS_NULL = ["(ทุกระดับ)"] + LEVELS
CATEGORIES = ["ข้าราชการ", "พนักงานราชการ", "ลูกจ้างประจำ"]
VACANCY_RULES = [
    "transfer_request | หนังสือขอรับโอน",
    "eligibility_list | ประกาศขึ้นบัญชี",
    "ready_to_fill | อัตราพร้อมบรรจุ",
]
CATEGORY_CODES = {
    "ข้าราชการ": "civil_servant",
    "พนักงานราชการ": "government_employee",
    "ลูกจ้างประจำ": "permanent_employee",
}
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
SERIAL_EPOCH = date(1899, 12, 30)  # Excel date serial เริ่มนับจากวันนี้

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TEMPLATE = ROOT / "docs" / "templates" / "P4_ข้อมูลงบบุคลากร_2569.xlsx"

SHEET_ALLOWANCE = "1_อัตราเงินเพิ่ม"
SHEET_SALARY = "2_อัตราเงินเดือน"
SHEET_POLICY = "3_นโยบายปี2569"
DATA_START_ROW = 3  # แถว 1 = หมายเหตุ · แถว 2 = header


@dataclass
class AllowanceRow:
    type_code: str
    level_code: str | None
    line_code: str | None
    amount: float | None
    percent: float | None
    derives_from_code: str | None
    fallback_amount: float | None
    effective_from: str
    effective_to: str | None
    doc_no: str
    sheet_row: int


@dataclass
class SalaryRow:
    category_code: str
    level_code: str
    min_amount: float
    max_amount: float
    effective_from: str
    doc_no: str
    sheet_row: int


@dataclass
class PolicyRow:
    vacancy_rule: str | None
    buffer_percent: float | None


@dataclass
class ImportResult:
    allowance_rows: list[AllowanceRow] = field(default_factory=list)
    salary_rows: list[SalaryRow] = field(default_factory=list)
    policy: PolicyRow | None = None
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _cell(v):
    """Normalize a cell value: strip strings, keep numbers, handle Excel dates."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _num(v, sheet_row, col, errors, sheet=SHEET_ALLOWANCE, decimals=2):
    """แปลงเป็น float — ปัดทศนิยมตามคอลัมน์เพื่อให้ dedup ตรงกับ DECIMAL ใน DB (F3)"""
    v = _cell(v)
    if v is None or v == "":
        return None
    try:
        if isinstance(v, (int, float)):
            f = float(v)
        else:
            f = float(str(v).replace(",", ""))
    except ValueError:
        errors.append(f"{sheet} แถว {sheet_row}: คอลัมน์ {col} ต้องเป็นตัวเลข ('{v}')")
        return None
    if f != f or f in (float("inf"), float("-inf")):
        errors.append(f"{sheet} แถว {sheet_row}: คอลัมน์ {col} ค่าตัวเลขไม่ถูกต้อง ('{v}')")
        return None
    return round(f, decimals)


def _date(v, sheet_row, col, errors, required=True, sheet=SHEET_ALLOWANCE):
    v = _cell(v)
    if v is None or v == "":
        if required:
            errors.append(f"{sheet} แถว {sheet_row}: คอลัมน์ {col} ต้องกรอกวันที่ (ตัวอย่าง: 2025-10-01)")
        return None
    if isinstance(v, str) and DATE_RE.match(v):
        return v
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        # Excel date serial — ตัวอย่าง: 45935 = 2025-10-01
        try:
            return (SERIAL_EPOCH + timedelta(days=float(v))).strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            pass
    errors.append(f"{sheet} แถว {sheet_row}: คอลัมน์ {col} วันที่ไม่ถูกต้อง ('{v}') — ใช้รูปแบบ YYYY-MM-DD เช่น 2025-10-01")
    return None


def _split_code_label(v):
    """'CODE | ชื่อไทย' -> CODE · ค่า '-', ว่าง หรือ '(ทุกระดับ)' -> None"""
    if v is None:
        return None
    if not isinstance(v, str):
        return None  # ตัวเลข/ค่าอื่น = ไม่ใช่ dropdown label (กัน crash + กันข้อมูลผิด)
    if v in ("-", "–"):
        return None
    return v.split(" | ", 1)[0].strip() if " | " in v else v.strip()


def _parse_allowance_sheet(ws, result: ImportResult, allowance_codes: set[str]):
    for r, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, max_col=11), start=DATA_START_ROW):
        vals = [_cell(c.value) for c in row]
        if all(v is None or v == "" for v in vals):
            continue  # แถวว่าง

        type_label = vals[0]
        if type_label is None:
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: ชนิดเงินเพิ่ม (คอลัมน์ A) ต้องกรอก")
            continue
        type_code = _split_code_label(type_label)
        if type_code is None or type_code not in allowance_codes:
            result.errors.append(
                f"{SHEET_ALLOWANCE} แถว {r}: ชนิดเงินเพิ่ม '{type_label}' ไม่ตรงกับรายการใน dropdown"
            )
            continue

        level_raw = vals[1]
        if level_raw is not None and not isinstance(level_raw, str):
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: ระดับตำแหน่ง (คอลัมน์ B) ต้องเป็นข้อความจาก dropdown")
            continue
        level = _split_code_label(level_raw)
        if level == "(ทุกระดับ)":
            level = None
        elif level is not None and level not in LEVELS:
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: ระดับตำแหน่ง '{level}' ไม่ตรงกับรายการใน dropdown")
            continue

        line_raw = vals[2]
        if line_raw is not None and not isinstance(line_raw, str):
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: สายงาน (คอลัมน์ C) ต้องเป็นข้อความ")
            continue
        line = _split_code_label(line_raw)
        amount = _num(vals[3], r, "D (จำนวนเงิน)", result.errors, sheet=SHEET_ALLOWANCE)
        percent = _num(vals[4], r, "E (ร้อยละ)", result.errors, sheet=SHEET_ALLOWANCE, decimals=3)
        derives_raw = vals[5]
        if derives_raw is not None and not isinstance(derives_raw, str):
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: 'เท่ากับชนิดเงินเพิ่ม' (คอลัมน์ F) ต้องเป็นข้อความจาก dropdown")
            continue
        derives = _split_code_label(derives_raw)
        fallback = _num(vals[6], r, "G (พื้นสำรอง)", result.errors, sheet=SHEET_ALLOWANCE)
        eff_from = _date(vals[7], r, "H (วันเริ่มมีผล)", result.errors)
        eff_to = _date(vals[8], r, "I (วันสิ้นสุด)", result.errors, required=False)
        doc_no = vals[9]
        if eff_from is None:
            continue  # error ถูกเพิ่มแล้ว — ไม่นำแถวที่มีวันเริ่มเป็น None ต่อไป
        if doc_no is None or isinstance(doc_no, (int, float)):
            if doc_no is None:
                result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: เลขที่คำสั่ง/ประกาศ (คอลัมน์ J) ต้องกรอก")
            else:
                result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: เลขที่คำสั่ง/ประกาศ (คอลัมน์ J) ต้องเป็นข้อความ")
            continue

        if derives is not None and derives not in allowance_codes:
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: 'เท่ากับชนิดเงินเพิ่ม' ('{derives}') ไม่ตรงกับรายการใน dropdown")
            continue
        if amount is None and percent is None and derives is None:
            result.errors.append(
                f"{SHEET_ALLOWANCE} แถว {r}: ต้องกรอกอย่างใดอย่างหนึ่ง — จำนวนเงิน (D) หรือ ร้อยละ (E) หรือ เท่ากับชนิดเงินเพิ่ม (F)"
            )
            continue
        # กติกาเดียวกับ CreateAllowanceRateDto::validate() — กันข้อมูลที่ระบบจะ reject อยู่ดี
        if amount is not None and percent is not None:
            result.errors.append(
                f"{SHEET_ALLOWANCE} แถว {r}: ใส่ได้ทีละอย่าง — จำนวนเงิน (D) หรือ ร้อยละ (E)"
            )
            continue
        if derives is not None and (amount is not None or percent is not None):
            result.errors.append(
                f"{SHEET_ALLOWANCE} แถว {r}: 'เท่ากับชนิดเงินเพิ่ม' (F) แล้ว — ห้ามใส่จำนวนเงิน/ร้อยละ (D/E)"
            )
            continue
        if derives is None and fallback is not None:
            result.errors.append(
                f"{SHEET_ALLOWANCE} แถว {r}: พื้นสำรอง (G) ใช้ได้เฉพาะกับ 'เท่ากับชนิดเงินเพิ่ม' (F)"
            )
            continue
        # ตรงกับ CreateAllowanceRateDto::validate() — ขอบเขตค่า
        if amount is not None and amount < 0:
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: จำนวนเงิน (D) ต้องไม่ติดลบ")
            continue
        if percent is not None and not (0 <= percent <= 100):
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: ร้อยละ (E) ต้องอยู่ระหว่าง 0–100")
            continue
        if fallback is not None and fallback < 0:
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: พื้นสำรอง (G) ต้องไม่ติดลบ")
            continue
        if eff_from is not None and eff_to is not None and eff_to < eff_from:
            result.errors.append(
                f"{SHEET_ALLOWANCE} แถว {r}: วันสิ้นสุด (I) ต้องไม่ก่อนวันเริ่มมีผล (H)"
            )
            continue
        # ความยาวตรงกับ schema (VARCHAR) — เกิน = DB error = rollback ทั้งไฟล์
        if line is not None and len(line) > 20:
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: สายงาน (C) ยาวเกิน 20 ตัวอักษร")
            continue
        if doc_no is not None and len(doc_no) > 100:
            result.errors.append(f"{SHEET_ALLOWANCE} แถว {r}: เลขที่คำสั่ง/ประกาศ (J) ยาวเกิน 100 ตัวอักษร")
            continue

        result.allowance_rows.append(
            AllowanceRow(
                type_code=type_code,
                level_code=level,
                line_code=line,
                amount=amount,
                percent=percent,
                derives_from_code=derives,
                fallback_amount=fallback if derives is not None else None,
                effective_from=eff_from,
                effective_to=eff_to,
                doc_no=doc_no,
                sheet_row=r,
            )
        )


def _parse_salary_sheet(ws, result: ImportResult):
    seen = {}  # (category, level, effective_from) -> แถวแรกที่เจอ
    for r, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, max_col=8), start=DATA_START_ROW):
        vals = [_cell(c.value) for c in row]
        if all(v is None or v == "" for v in vals):
            continue

        category_label = vals[0]
        if category_label is None:
            result.errors.append(f"{SHEET_SALARY} แถว {r}: ประเภทบุคลากร (คอลัมน์ A) ต้องกรอก")
            continue
        category_code = CATEGORY_CODES.get(category_label)
        if category_code is None:
            result.errors.append(f"{SHEET_SALARY} แถว {r}: ประเภทบุคลากร '{category_label}' ไม่ตรงกับรายการใน dropdown")
            continue

        level = _cell(vals[2])
        if level is None:
            result.errors.append(f"{SHEET_SALARY} แถว {r}: ระดับตำแหน่ง (คอลัมน์ C) ต้องกรอก")
            continue
        if level not in LEVELS:
            result.errors.append(f"{SHEET_SALARY} แถว {r}: ระดับตำแหน่ง '{level}' ไม่ตรงกับรายการใน dropdown")
            continue

        min_amount = _num(vals[3], r, "D (ขั้นต่ำ)", result.errors, sheet=SHEET_SALARY)
        max_amount = _num(vals[4], r, "E (ขั้นสูง)", result.errors, sheet=SHEET_SALARY)
        if min_amount is None or max_amount is None:
            result.errors.append(f"{SHEET_SALARY} แถว {r}: เงินเดือนขั้นต่ำ (D) และขั้นสูง (E) ต้องกรอกทั้งคู่")
            continue
        if max_amount < min_amount:
            result.errors.append(f"{SHEET_SALARY} แถว {r}: เงินเดือนขั้นสูง ({max_amount:g}) ต่ำกว่าขั้นต่ำ ({min_amount:g})")
            continue

        eff_from = _date(vals[5], r, "F (วันเริ่มใช้)", result.errors, sheet=SHEET_SALARY)
        doc_no = vals[6]
        if eff_from is None:
            continue  # error ถูกเพิ่มแล้ว — ไม่นำแถวที่มีวันเริ่มเป็น None ต่อไป
        if doc_no is None or isinstance(doc_no, (int, float)):
            if doc_no is None:
                result.errors.append(f"{SHEET_SALARY} แถว {r}: เลขที่คำสั่ง/ประกาศ (คอลัมน์ G) ต้องกรอก")
            else:
                result.errors.append(f"{SHEET_SALARY} แถว {r}: เลขที่คำสั่ง/ประกาศ (คอลัมน์ G) ต้องเป็นข้อความ")
            continue

        # UK (employee_category, level_code, effective_from) — ซ้ำในไฟล์ = UPSERT ทับกันเงียบๆ
        key = (category_code, level, eff_from)
        if key in seen:
            result.errors.append(
                f"{SHEET_SALARY} แถว {r}: ซ้ำกับแถว {seen[key]} — ({category_label}, {level}, {eff_from}) "
                "มีแถวเดียวได้ต่อ (ประเภท, ระดับ, วันเริ่มใช้) ชุดเดียวกัน"
            )
            continue
        seen[key] = r

        result.salary_rows.append(
            SalaryRow(
                category_code=category_code,
                level_code=level,
                min_amount=min_amount,
                max_amount=max_amount,
                effective_from=eff_from,
                doc_no=doc_no,
                sheet_row=r,
            )
        )


def _parse_policy_sheet(ws, result: ImportResult):
    """หาค่าตามชื่อรายการในคอลัมน์ A — ไม่ผูกแถวตายตัว (HR อาจแทรก/ลบแถว)"""
    values = {}
    for r, row in enumerate(ws.iter_rows(min_row=2, max_col=2), start=2):
        label = _cell(row[0].value)
        if label:
            values[label.strip()] = _cell(row[1].value)

    vacancy = None
    v = values.get("เกณฑ์อัตราว่างปี 2569 (vacancy_rule)")
    if v not in (None, "", "-"):
        code = _split_code_label(v)
        if code not in [r.split(" | ", 1)[0] for r in VACANCY_RULES]:
            result.errors.append(f"{SHEET_POLICY}: เกณฑ์อัตราว่าง '{v}' ไม่ตรงกับรายการใน dropdown")
        else:
            vacancy = code

    buffer_pct = None
    b = values.get("ช่องปรับ buffer (%)")
    if b not in (None, "", "-"):
        try:
            buffer_pct = float(str(b).replace("%", "").replace(",", ""))
        except ValueError:
            result.errors.append(f"{SHEET_POLICY}: ช่องปรับ buffer (%) ต้องเป็นตัวเลข ('{b}')")

    result.policy = PolicyRow(vacancy_rule=vacancy, buffer_percent=buffer_pct)


def parse_workbook(path: str | Path) -> ImportResult:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return _parse_workbook_loaded(wb)
    finally:
        wb.close()


def _parse_workbook_loaded(wb) -> ImportResult:
    result = ImportResult()

    # ค่าที่ใช้ได้จาก dropdown ชนิดเงินเพิ่ม — มาจาก Lists sheet ในไฟล์เอง
    allowance_codes = set()
    if "Lists" not in wb.sheetnames:
        result.errors.append("ไม่พบแผ่น 'Lists' (แหล่ง dropdown) — ใช้ไฟล์ที่ได้จาก build_p4_template.py")
        return result
    ws_lists = wb["Lists"]
    for row in ws_lists.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                code = _split_code_label(_cell(cell.value))
                if code and code != "-":
                    allowance_codes.add(code)

    for name in (SHEET_ALLOWANCE, SHEET_SALARY):
        if name not in wb.sheetnames:
            result.errors.append(f"ไม่พบแผ่น '{name}' ในไฟล์ — ใช้ไฟล์ที่ได้จาก build_p4_template.py")
            return result
    if SHEET_POLICY not in wb.sheetnames:
        result.errors.append(f"ไม่พบแผ่น '{SHEET_POLICY}' ในไฟล์")
        return result

    _parse_allowance_sheet(wb[SHEET_ALLOWANCE], result, allowance_codes)
    _parse_salary_sheet(wb[SHEET_SALARY], result)
    _parse_policy_sheet(wb[SHEET_POLICY], result)
    return result


# ---- ส่วนต่อ DB -------------------------------------------------------------
def resolve_type_ids(conn, codes: set[str]) -> dict[str, int]:
    """code -> allowance_types.id (อ่านสดจาก DB กัน drift)"""
    if not codes:
        return {}
    placeholders = ",".join(["%s"] * len(codes))
    cur = conn.cursor(dictionary=True)
    cur.execute(
        f"SELECT code, id FROM allowance_types WHERE code IN ({placeholders}) "
        "AND is_active = 1 AND deleted_at IS NULL",
        sorted(codes),
    )
    return {r["code"]: r["id"] for r in cur.fetchall()}


def _insert_allowance_rows(conn, rows: list[AllowanceRow], type_ids: dict[str, int]):
    """INSERT ใหม่ทั้งหมด — แถวซ้ำ (ทุกคอลัมน์เหมือนเดิม) ข้ามเงียบๆ"""
    cur = conn.cursor()
    inserted = skipped = 0
    for r in rows:
        type_id = type_ids[r.type_code]
        derives_id = type_ids.get(r.derives_from_code) if r.derives_from_code else None
        cur.execute(
            "SELECT COUNT(*) FROM allowance_rates "
            "WHERE allowance_type_id = %s AND level_code <=> %s AND line_code <=> %s "
            "AND amount <=> %s AND percent <=> %s AND derives_from_type_id <=> %s "
            "AND fallback_amount <=> %s AND effective_from = %s AND effective_to <=> %s "
            "AND doc_no <=> %s AND deleted_at IS NULL",
            (type_id, r.level_code, r.line_code, r.amount, r.percent, derives_id,
             r.fallback_amount, r.effective_from, r.effective_to, r.doc_no),
        )
        if cur.fetchone()[0] > 0:
            skipped += 1
            continue
        cur.execute(
            "INSERT INTO allowance_rates "
            "(allowance_type_id, level_code, line_code, amount, percent, derives_from_type_id, "
            " fallback_amount, effective_from, effective_to, doc_no) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (type_id, r.level_code, r.line_code, r.amount, r.percent, derives_id,
             r.fallback_amount, r.effective_from, r.effective_to, r.doc_no),
        )
        inserted += 1
    return inserted, skipped


def _insert_salary_rows(conn, rows: list[SalaryRow]):
    """UPSERT ตาม UK (employee_category, level_code, effective_from) — import ซ้ำได้

    กันแถว soft-deleted: UK ไม่รวม deleted_at — ถ้าแถวเดิมถูก soft-delete แล้ว
    UPSERT จะฟื้นแถวที่ตั้งใจลบขึ้นมาเงียบๆ ⇒ เจอแล้ว error (ต้องจัดการด้วยมือ)
    """
    cur = conn.cursor()
    inserted = updated = 0
    for r in rows:
        cur.execute(
            "SELECT deleted_at FROM salary_scales "
            "WHERE employee_category = %s AND level_code = %s AND effective_from = %s",
            (r.category_code, r.level_code, r.effective_from),
        )
        existing = cur.fetchone()
        if existing is not None and existing[0] is not None:
            raise RuntimeError(
                f"salary_scales แถว ({r.category_code}, {r.level_code}, {r.effective_from}) "
                "ถูกยกเลิกในระบบแล้ว (soft-deleted) — จัดการด้วยมือก่อน import ซ้ำ"
            )
        cur.execute(
            "INSERT INTO salary_scales "
            "(employee_category, level_code, min_amount, max_amount, effective_from, doc_no) "
            "VALUES (%s, %s, %s, %s, %s, %s) AS new "
            "ON DUPLICATE KEY UPDATE min_amount = new.min_amount, "
            "max_amount = new.max_amount, doc_no = new.doc_no, "
            "deleted_at = NULL, is_active = 1",
            (r.category_code, r.level_code, r.min_amount, r.max_amount, r.effective_from, r.doc_no),
        )
        if cur.rowcount == 1:
            inserted += 1
        elif cur.rowcount == 2:
            updated += 1
    return inserted, updated


def _update_policy(conn, policy: PolicyRow, fiscal_year_id: int):
    if policy is None or (policy.vacancy_rule is None and policy.buffer_percent is None):
        return None
    sets, params = [], []
    if policy.vacancy_rule is not None:
        sets.append("vacancy_rule = %s")
        params.append(policy.vacancy_rule)
    if policy.buffer_percent is not None:
        sets.append("buffer_percent = %s")
        params.append(policy.buffer_percent)
    params.append(fiscal_year_id)
    cur = conn.cursor()
    cur.execute(
        f"UPDATE personnel_budget_policies SET {', '.join(sets)} WHERE fiscal_year_id = %s "
        "AND deleted_at IS NULL",
        params,
    )
    return cur.rowcount


def apply_import(conn, result: ImportResult, fiscal_year: int) -> dict:
    # รวม derives codes ด้วย — derives อาจอ้าง type ที่ไม่มีแถวในไฟล์นี้
    codes = {r.type_code for r in result.allowance_rows}
    codes |= {r.derives_from_code for r in result.allowance_rows if r.derives_from_code}
    type_ids = resolve_type_ids(conn, codes)
    missing = codes - set(type_ids)
    if missing:
        raise RuntimeError(f"allowance type code(s) not in DB: {sorted(missing)}")

    cur = conn.cursor()
    cur.execute("SELECT id FROM fiscal_years WHERE year = %s AND deleted_at IS NULL", (fiscal_year,))
    fy_row = cur.fetchone()
    if fy_row is None:
        raise RuntimeError(f"fiscal year {fiscal_year} not found in DB")

    a_ins, a_skip = _insert_allowance_rows(conn, result.allowance_rows, type_ids)
    s_ins, s_upd = _insert_salary_rows(conn, result.salary_rows)
    p_upd = _update_policy(conn, result.policy, fy_row[0])
    return {
        "allowance_rates": {"inserted": a_ins, "skipped_duplicates": a_skip},
        "salary_scales": {"inserted": s_ins, "updated": s_upd},
        "policy_updated": p_upd,
    }


def main(argv=None):
    parser = argparse.ArgumentParser(description="นำเข้าเทมเพลต P4 (ข้อมูลงบบุคลากร 2569)")
    parser.add_argument("path", nargs="?", default=str(DEFAULT_TEMPLATE),
                        help="ไฟล์ xlsx จาก HR (ค่าเริ่มต้น: เทมเพลตใน docs/templates)")
    parser.add_argument("--apply", action="store_true", help="นำเข้าจริง (ค่าเริ่มต้น = dry run เท่านั้น)")
    parser.add_argument("--fiscal-year", type=int, default=2569, help="ปีงบประมาณ (ค่าเริ่มต้น: 2569)")
    args = parser.parse_args(argv)

    if not Path(args.path).exists():
        print(f"ไม่พบไฟล์: {args.path}")
        return 2

    result = parse_workbook(args.path)

    print("=" * 60)
    print("ผลตรวจเทมเพลต P4 — {}".format("DRY RUN (ยังไม่เขียน DB)" if not args.apply else "นำเข้าจริง"))
    print("=" * 60)
    print(f"อัตราเงินเพิ่ม : {len(result.allowance_rows)} แถว")
    print(f"อัตราเงินเดือน : {len(result.salary_rows)} แถว")
    for w in result.warnings:
        print(f"  ⚠ {w}")
    if result.errors:
        print()
        print(f"พบ {len(result.errors)} ข้อผิดพลาด — แก้ในไฟล์แล้วรันใหม่:")
        for e in result.errors:
            print(f"  ✗ {e}")
        print()
        print("ยังไม่มีการเขียนข้อมูลใดๆ ลงฐานข้อมูล")
        return 1

    if not args.apply:
        for r in result.allowance_rows:
            detail = r.amount if r.amount is not None else (f"{r.percent}%" if r.percent is not None else f"= {r.derives_from_code}")
            print(f"  + [{r.type_code}] level={r.level_code or '*'} line={r.line_code or '*'} "
                  f"{detail} จาก {r.effective_from} doc={r.doc_no}")
        for r in result.salary_rows:
            print(f"  + [{r.category_code}] {r.level_code} {r.min_amount:g}–{r.max_amount:g} "
                  f"จาก {r.effective_from} doc={r.doc_no}")
        if result.policy:
            print(f"  + นโยบาย: vacancy_rule={result.policy.vacancy_rule or '(ไม่เปลี่ยน)'} "
                  f"buffer={result.policy.buffer_percent or '(ไม่เปลี่ยน)'}")
        print()
        print("ผ่านตรวจสอบ — รันด้วย --apply เพื่อนำเข้าจริง")
        return 0

    # กันรัน --apply พลาดโดยไม่ตั้งใจ
    try:
        print()
        answer = input("ยืนยันนำเข้าจริง (เขียนลงฐานข้อมูล) หรือไม่? [y/N]: ").strip().lower()
    except EOFError:
        print("ไม่มี stdin — ยกเลิก (รันด้วยพารามิเตอร์ให้ชัดเจน)")
        return 1
    if answer != "y":
        print("ยกเลิก — ไม่ได้เขียนข้อมูลใดๆ")
        return 0

    try:
        from db_config import get_db_config
        import mysql.connector
        cfg = get_db_config()
        conn = mysql.connector.connect(
            host=cfg["host"], port=cfg["port"], user=cfg["user"],
            password=cfg["password"], database=cfg["database"], charset="utf8mb4",
        )
    except mysql.connector.Error as e:
        print(f"ต่อฐานข้อมูลไม่ได้ — เปิด MySQL ก่อน: {e}")
        return 2

    try:
        conn.start_transaction()
        stats = apply_import(conn, result, args.fiscal_year)
        conn.commit()
    except Exception as e:  # noqa: BLE001
        conn.rollback()
        print(f"นำเข้าไม่สำเร็จ — ยกเลิกทั้งหมด (rollback): {e}")
        return 1
    finally:
        conn.close()

    print()
    print("นำเข้าสำเร็จ (ปีงบ " + str(args.fiscal_year) + "):")
    print(f"  allowance_rates : เพิ่ม {stats['allowance_rates']['inserted']} "
          f"· ข้ามซ้ำ {stats['allowance_rates']['skipped_duplicates']}")
    print(f"  salary_scales   : เพิ่ม {stats['salary_scales']['inserted']} "
          f"· อัปเดต {stats['salary_scales']['updated']}")
    print(f"  นโยบาย          : อัปเดต {stats['policy_updated']} แถว")
    return 0


if __name__ == "__main__":
    sys.exit(main())
