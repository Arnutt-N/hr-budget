#!/usr/bin/env python3
"""Tests for import_p4_template.py — parse/validate only, no DB writes.

Run (from python/): venv/Scripts/python.exe -m unittest test_import_p4_template -v
"""
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from import_p4_template import (  # noqa: E402
    CATEGORY_CODES,
    SHEET_ALLOWANCE,
    SHEET_POLICY,
    SHEET_SALARY,
    SalaryRow,
    _insert_salary_rows,
    parse_workbook,
)


class FakeCursor:
    """เลียนแบบ cursor มาตรฐาน (execute/fetchone/rowcount) สำหรับ test DB-layer"""

    def __init__(self, select_results):
        self._select_results = select_results  # ผล SELECT ที่จะตอบตามลำดับ
        self.rowcount = 0

    def execute(self, sql, params=None):
        if sql.startswith("SELECT"):
            self._result = self._select_results.pop(0)
        else:
            self._result = None
            self.rowcount = 1

    def fetchone(self):
        return self._result


class FakeConn:
    def __init__(self, select_results):
        self._select_results = select_results

    def cursor(self):
        return FakeCursor(self._select_results)


def build_workbook(allowance_rows=None, salary_rows=None, vacancy=None, buffer_pct=None):
    """Build a minimal workbook with the same sheet names/layout as the template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lists"
    lists = [
        ("ชนิดเงินเพิ่ม", ["POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "KHN | ค.ต.น.", "HOUSE_RENT | ค่าเช่าบ้าน"]),
        ("ระดับตำแหน่ง (รวมทุกระดับ)", ["(ทุกระดับ)", "ปฏิบัติงาน", "ชำนาญงาน", "อาวุโส", "ทักษะพิเศษ"]),
        ("ระดับตำแหน่ง", ["ปฏิบัติงาน", "ชำนาญงาน", "อาวุโส", "ทักษะพิเศษ"]),
        ("ประเภทบุคลากร", ["ข้าราชการ", "พนักงานราชการ", "ลูกจ้างประจำ"]),
    ]
    for i, (name, values) in enumerate(lists, start=1):
        ws.cell(row=1, column=i, value=name)
        for r, v in enumerate(values, start=2):
            ws.cell(row=r, column=i, value=v)

    ws1 = wb.create_sheet(SHEET_ALLOWANCE)
    ws1.cell(row=2, column=1, value="ชนิดเงินเพิ่ม *")
    for r, row in enumerate(allowance_rows or [], start=3):
        for c, v in enumerate(row, start=1):
            ws1.cell(row=r, column=c, value=v)

    ws2 = wb.create_sheet(SHEET_SALARY)
    ws2.cell(row=2, column=1, value="ประเภทบุคลากร *")
    for r, row in enumerate(salary_rows or [], start=3):
        for c, v in enumerate(row, start=1):
            ws2.cell(row=r, column=c, value=v)

    ws3 = wb.create_sheet(SHEET_POLICY)
    policy_rows = [
        ("เกณฑ์อัตราว่างปี 2569 (vacancy_rule)", vacancy),
        ("เลขที่หนังสือเวียน/ประกาศ", ""),
        ("วันที่หนังสือ", ""),
        ("โหมดคำนวณ (calc_mode)", "prorate"),
        ("ช่องปรับ buffer (%)", buffer_pct),
    ]
    for i, (label, value) in enumerate(policy_rows, start=3):
        ws3.cell(row=i, column=1, value=label)
        ws3.cell(row=i, column=2, value=value)
    return wb


def parse(wb):
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "p4.xlsx"
        wb.save(p)
        return parse_workbook(p)


class ParseAllowanceSheetTest(unittest.TestCase):
    def test_valid_row_with_amount(self):
        res = parse(build_workbook(allowance_rows=[
            ["POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "(ทุกระดับ)", "", 4200, "", "-", "", "2025-10-01", "", "คำสั่ง 1/2569", ""],
        ]))
        self.assertEqual(res.errors, [])
        self.assertEqual(len(res.allowance_rows), 1)
        r = res.allowance_rows[0]
        self.assertEqual(r.type_code, "POSITION_ALLOWANCE")
        self.assertIsNone(r.level_code)
        self.assertEqual(r.amount, 4200.0)
        self.assertIsNone(r.percent)
        self.assertIsNone(r.derives_from_code)
        self.assertEqual(r.effective_from, "2025-10-01")

    def test_percent_row(self):
        res = parse(build_workbook(allowance_rows=[
            ["KHN | ค.ต.น.", "(ทุกระดับ)", "", "", "", "POSITION_ALLOWANCE | เงินประจำตำแหน่ง", 3500, "2025-10-01", "", "คำสั่ง 2/2569", ""],
        ]))
        self.assertEqual(res.errors, [])
        r = res.allowance_rows[0]
        self.assertIsNone(r.percent)
        self.assertEqual(r.derives_from_code, "POSITION_ALLOWANCE")
        self.assertEqual(r.fallback_amount, 3500.0)

    def test_amount_row_with_percent_only_type(self):
        res = parse(build_workbook(allowance_rows=[
            ["KHN | ค.ต.น.", "(ทุกระดับ)", "", "", 10.0, "-", "", "2025-10-01", "", "คำสั่ง 15/2569", ""],
        ]))
        self.assertEqual(res.errors, [])
        r = res.allowance_rows[0]
        self.assertEqual(r.percent, 10.0)
        self.assertIsNone(r.derives_from_code)

    def test_missing_type_code_rejected(self):
        res = parse(build_workbook(allowance_rows=[
            ["", "(ทุกระดับ)", "", 4200, "", "-", "", "2025-10-01", "", "คำสั่ง 3/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ชนิดเงินเพิ่ม", res.errors[0])
        self.assertEqual(res.allowance_rows, [])

    def test_unknown_type_rejected(self):
        res = parse(build_workbook(allowance_rows=[
            ["NOPE | ไม่มีในระบบ", "(ทุกระดับ)", "", 4200, "", "-", "", "2025-10-01", "", "คำสั่ง 4/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ไม่ตรงกับรายการ", res.errors[0])

    def test_no_amount_percent_or_derived_rejected(self):
        res = parse(build_workbook(allowance_rows=[
            ["POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "(ทุกระดับ)", "", "", "", "-", "", "2025-10-01", "", "คำสั่ง 5/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ต้องกรอกอย่างใดอย่างหนึ่ง", res.errors[0])

    def test_amount_and_percent_both_rejected(self):
        """กติกาเดียวกับ CreateAllowanceRateDto::validate() — ใส่ได้ทีละอย่าง"""
        res = parse(build_workbook(allowance_rows=[
            ["POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "(ทุกระดับ)", "", 4200, 10.0, "-", "", "2025-10-01", "", "คำสั่ง 12/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ทีละอย่าง", res.errors[0])
        self.assertEqual(res.allowance_rows, [])

    def test_derived_with_amount_rejected(self):
        res = parse(build_workbook(allowance_rows=[
            ["KHN | ค.ต.น.", "(ทุกระดับ)", "", 4200, "", "POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "", "2025-10-01", "", "คำสั่ง 13/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ห้ามใส่", res.errors[0])
        self.assertEqual(res.allowance_rows, [])

    def test_fallback_without_derived_rejected(self):
        res = parse(build_workbook(allowance_rows=[
            ["POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "(ทุกระดับ)", "", 4200, "", "-", 3500, "2025-10-01", "", "คำสั่ง 14/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("พื้นสำรอง", res.errors[0])
        self.assertEqual(res.allowance_rows, [])

    def test_invalid_date_rejected(self):
        res = parse(build_workbook(allowance_rows=[
            ["POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "(ทุกระดับ)", "", 4200, "", "-", "", "01/10/2025", "", "คำสั่ง 6/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("วันที่ไม่ถูกต้อง", res.errors[0])

    def test_excel_date_serial_accepted(self):
        """Excel จัดเก็บวันที่เป็น serial number (45931 = 2025-10-01) — ต้องแปลงได้"""
        res = parse(build_workbook(allowance_rows=[
            ["POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "(ทุกระดับ)", "", 4200, "", "-", "", 45931, "", "คำสั่ง 16/2569", ""],
        ]))
        self.assertEqual(res.errors, [])
        self.assertEqual(res.allowance_rows[0].effective_from, "2025-10-01")

    def test_duplicate_row_blank_lines_skipped(self):
        res = parse(build_workbook(allowance_rows=[
            ["POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "(ทุกระดับ)", "", 4200, "", "-", "", "2025-10-01", "", "คำสั่ง 7/2569", ""],
            [None] * 11,
        ]))
        self.assertEqual(res.errors, [])
        self.assertEqual(len(res.allowance_rows), 1)


class ParseSalarySheetTest(unittest.TestCase):
    def test_valid_row(self):
        res = parse(build_workbook(salary_rows=[
            ["ข้าราชการ", "ประเภททั่วไป", "ปฏิบัติงาน", 15000, 25000, "2025-10-01", "คำสั่ง 8/2569", ""],
        ]))
        self.assertEqual(res.errors, [])
        self.assertEqual(len(res.salary_rows), 1)
        r = res.salary_rows[0]
        self.assertEqual(r.category_code, "civil_servant")
        self.assertEqual(r.level_code, "ปฏิบัติงาน")
        self.assertEqual(r.min_amount, 15000.0)
        self.assertEqual(r.max_amount, 25000.0)

    def test_category_label_mapping(self):
        self.assertEqual(CATEGORY_CODES["ข้าราชการ"], "civil_servant")
        self.assertEqual(CATEGORY_CODES["พนักงานราชการ"], "government_employee")
        self.assertEqual(CATEGORY_CODES["ลูกจ้างประจำ"], "permanent_employee")

    def test_unknown_category_rejected(self):
        res = parse(build_workbook(salary_rows=[
            ["ข้าราชการสนาม", "ประเภททั่วไป", "ปฏิบัติงาน", 15000, 25000, "2025-10-01", "คำสั่ง 9/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ประเภทบุคลากร", res.errors[0])

    def test_max_below_min_rejected(self):
        res = parse(build_workbook(salary_rows=[
            ["ข้าราชการ", "ประเภททั่วไป", "ปฏิบัติงาน", 25000, 15000, "2025-10-01", "คำสั่ง 10/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("สูง", res.errors[0])

    def test_unknown_level_rejected(self):
        res = parse(build_workbook(salary_rows=[
            ["ข้าราชการ", "ประเภททั่วไป", "ชำนาญการพิเศษ", 15000, 25000, "2025-10-01", "คำสั่ง 11/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ระดับตำแหน่ง", res.errors[0])

    def test_missing_min_or_max_rejected(self):
        res = parse(build_workbook(salary_rows=[
            ["ข้าราชการ", "ประเภททั่วไป", "ปฏิบัติงาน", "", 25000, "2025-10-01", "คำสั่ง 17/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ขั้นต่ำ (D) และขั้นสูง (E)", res.errors[0])
        self.assertEqual(res.salary_rows, [])

    def test_duplicate_uk_key_rejected(self):
        """UK (employee_category, level_code, effective_from) — ซ้ำในไฟล์ = UPSERT ทับกันเงียบๆ"""
        res = parse(build_workbook(salary_rows=[
            ["ข้าราชการ", "ประเภททั่วไป", "ปฏิบัติงาน", 15000, 25000, "2025-10-01", "คำสั่ง 18/2569", ""],
            ["ข้าราชการ", "ประเภททั่วไป", "ปฏิบัติงาน", 16000, 26000, "2025-10-01", "คำสั่ง 19/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ซ้ำกับแถว 3", res.errors[0])
        self.assertEqual(len(res.salary_rows), 1)


class ParsePolicySheetTest(unittest.TestCase):
    def test_valid_vacancy(self):
        res = parse(build_workbook(vacancy="eligibility_list | ประกาศขึ้นบัญชี"))
        self.assertEqual(res.errors, [])
        self.assertEqual(res.policy.vacancy_rule, "eligibility_list")

    def test_empty_vacancy_ok(self):
        res = parse(build_workbook(vacancy=None))
        self.assertEqual(res.errors, [])
        self.assertIsNone(res.policy.vacancy_rule)

    def test_invalid_vacancy_rejected(self):
        res = parse(build_workbook(vacancy="bogus | ผิด"))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("เกณฑ์อัตราว่าง", res.errors[0])

    def test_buffer_percent_parsed(self):
        res = parse(build_workbook(buffer_pct="2.5"))
        self.assertEqual(res.policy.buffer_percent, 2.5)

    def test_invalid_buffer_rejected(self):
        res = parse(build_workbook(buffer_pct="abc"))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("buffer", res.errors[0])

    def test_policy_parsed_by_label_not_fixed_row(self):
        """หาค่าจากชื่อรายการ (คอลัมน์ A) — ถ้า HR แทรกแถวกลาง แถวเลื่อนก็ยังอ่านถูก"""
        wb = build_workbook()
        ws3 = wb[SHEET_POLICY]
        # ลบแถว 3 และ 7 ทิ้ง (จำลอง HR ลบ/เพิ่มแถว) แล้วเขียนใหม่แถว 4 และ 8
        for r, row in enumerate([["หมายเหตุเพิ่ม", "เพิ่มเอง", "xxx"],
                                 ["เกณฑ์อัตราว่างปี 2569 (vacancy_rule)", "transfer_request | หนังสือขอรับโอน", ""],
                                 ["เลขที่หนังสือเวียน/ประกาศ", "ว 123/2569", ""],
                                 ["วันที่หนังสือ", "2025-09-01", ""],
                                 ["โหมดคำนวณ (calc_mode)", "prorate", ""],
                                 ["ช่องปรับ buffer (%)", "3.5", ""]], start=3):
            for c, v in enumerate(row, start=1):
                ws3.cell(row=r, column=c, value=v)
        res = parse(wb)
        self.assertEqual(res.errors, [])
        self.assertEqual(res.policy.vacancy_rule, "transfer_request")
        self.assertEqual(res.policy.buffer_percent, 3.5)

    def test_negative_amount_rejected(self):
        res = parse(build_workbook(allowance_rows=[
            ["POSITION_ALLOWANCE | เงินประจำตำแหน่ง", "(ทุกระดับ)", "", -100, "", "-", "", "2025-10-01", "", "คำสั่ง 20/2569", ""],
        ]))
        self.assertEqual(len(res.errors), 1)
        self.assertIn("ไม่ติดลบ", res.errors[0])
        self.assertEqual(res.allowance_rows, [])


class WorkbookShapeTest(unittest.TestCase):
    def test_missing_sheet_reported(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Lists"
        with tempfile.TemporaryDirectory() as d:
            p = Path(d) / "p4.xlsx"
            wb.save(p)
            res = parse_workbook(p)
        self.assertEqual(len(res.errors), 1)
        self.assertIn(SHEET_ALLOWANCE, res.errors[0])


class DbLayerTest(unittest.TestCase):
    def make_row(self):
        return SalaryRow(
            category_code="civil_servant",
            level_code="ปฏิบัติงาน",
            min_amount=15000.0,
            max_amount=25000.0,
            effective_from="2025-10-01",
            doc_no="คำสั่ง 99/2569",
            sheet_row=3,
        )

    def test_salary_upsert_blocks_soft_deleted_row(self):
        """แถว soft-deleted (deleted_at ไม่ NULL) ต้อง block — กันฟื้นแถวที่ตั้งใจลบ"""
        conn = FakeConn([("2025-01-01 10:00:00",)])  # SELECT คืน deleted_at ที่ไม่ NULL
        with self.assertRaises(RuntimeError) as ctx:
            _insert_salary_rows(conn, [self.make_row()])
        self.assertIn("soft-deleted", str(ctx.exception))

    def test_salary_upsert_allows_active_row(self):
        """แถว active (deleted_at NULL) → INSERT/UPSERT ได้ปกติ"""
        conn = FakeConn([(None,)])  # SELECT คืน deleted_at = NULL
        inserted, updated = _insert_salary_rows(conn, [self.make_row()])
        self.assertEqual((inserted, updated), (1, 0))


if __name__ == "__main__":
    unittest.main()
